import io
from datetime import time

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session

import models
from auth import get_current_user
from database import get_db

router = APIRouter(dependencies=[Depends(get_current_user)])

TRACK_BG = ["EDE9FE", "D1FAE5", "FFEDD5", "FCE7F3", "CCFBF1", "E0E7FF", "FFE4E6", "CFFAFE"]
TRACK_FG = ["5B21B6", "065F46", "9A3412", "9D174D", "134E4A", "3730A3", "9F1239", "164E63"]

SLOT_MIN = 20
GRID_START = 10 * 60
GRID_END = 18 * 60
TOTAL_SLOTS = (GRID_END - GRID_START) // SLOT_MIN


def _time_to_slot(t: time) -> int:
    return ((t.hour * 60 + t.minute) - GRID_START) // SLOT_MIN


def _slot_label(slot: int) -> str:
    m = GRID_START + slot * SLOT_MIN
    return f"{m // 60:02d}:{m % 60:02d}"


@router.get("/conferences/{conference_id}/schedule/export")
def export_schedule_excel(
    conference_id: int,
    version_id: int | None = Query(default=None),
    db: Session = Depends(get_db),
):
    import openpyxl
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    conference = db.get(models.Conference, conference_id)
    if not conference:
        raise HTTPException(status_code=404, detail="Conference not found")

    halls = sorted(conference.halls, key=lambda h: h.id)
    tracks = conference.tracks
    track_idx_map = {t.id: i for i, t in enumerate(tracks)}

    thin_side = Side(style="thin", color="E5E7EB")
    hour_side = Side(style="medium", color="9CA3AF")

    def thin_border():
        return Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    def hour_border():
        return Border(left=thin_side, right=thin_side, top=hour_side, bottom=thin_side)

    # Собираем доклады по дням из выбранной версии
    talks_by_day: dict[int, list[dict]] = {d.id: [] for d in conference.days}
    breaks_by_day: dict[int, list] = {d.id: list(d.breaks) for d in conference.days}

    if version_id is not None:
        version = db.get(models.ScheduleVersion, version_id)
        if not version or version.conference_id != conference_id:
            raise HTTPException(status_code=404, detail="Version not found")
        for p in version.placements:
            if p.day_id not in talks_by_day:
                continue
            t = p.talk
            ti = track_idx_map.get(t.primary_track_id, -1) if t.primary_track_id else -1
            talks_by_day[p.day_id].append({
                "title": t.title,
                "speaker": t.speaker_name or "",
                "track": t.primary_track.name if t.primary_track else "",
                "hall_id": p.hall_id,
                "start": p.start_time,
                "end": p.end_time,
                "track_idx": ti,
                "is_keynote": t.speaker_level == "keynote",
            })

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Расписание"

    n_halls = len(halls)
    total_cols = 1 + n_halls

    ws.column_dimensions["A"].width = 7
    for i in range(n_halls):
        ws.column_dimensions[get_column_letter(i + 2)].width = 26

    hall_col: dict[int, int] = {hall.id: i + 2 for i, hall in enumerate(halls)}
    current_row = 1
    sorted_days = sorted(conference.days, key=lambda d: d.date)

    for day_idx, day in enumerate(sorted_days):
        if n_halls == 0:
            continue
        day_talks = talks_by_day[day.id]
        day_breaks = breaks_by_day[day.id]

        DAY_HEADER_ROW = current_row
        HALL_HEADER_ROW = current_row + 1
        GRID_FIRST_ROW = current_row + 2

        ws.merge_cells(start_row=DAY_HEADER_ROW, start_column=1, end_row=DAY_HEADER_ROW, end_column=total_cols)
        c = ws.cell(row=DAY_HEADER_ROW, column=1)
        c.value = day.date.strftime("%A, %d %B %Y").capitalize()
        c.font = Font(bold=True, size=12, color="1F2937")
        c.fill = PatternFill("solid", fgColor="EEF2FF")
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[DAY_HEADER_ROW].height = 24

        tc = ws.cell(row=HALL_HEADER_ROW, column=1, value="Время")
        tc.font = Font(bold=True, color="FFFFFF", size=10)
        tc.fill = PatternFill("solid", fgColor="4F46E5")
        tc.alignment = Alignment(horizontal="center", vertical="center")
        tc.border = thin_border()
        ws.row_dimensions[HALL_HEADER_ROW].height = 32

        for i, hall in enumerate(halls):
            col = i + 2
            hc = ws.cell(row=HALL_HEADER_ROW, column=col, value=f"{hall.name}\n{hall.capacity} мест")
            hc.font = Font(bold=True, color="FFFFFF", size=10)
            hc.fill = PatternFill("solid", fgColor="4F46E5")
            hc.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            hc.border = thin_border()

        for slot in range(TOTAL_SLOTS):
            row = GRID_FIRST_ROW + slot
            is_hour = (GRID_START + slot * SLOT_MIN) % 60 == 0
            b = hour_border() if is_hour else thin_border()

            lc = ws.cell(row=row, column=1, value=_slot_label(slot))
            lc.font = Font(size=8, bold=is_hour, color="374151" if is_hour else "9CA3AF")
            lc.alignment = Alignment(horizontal="right", vertical="top")
            lc.fill = PatternFill("solid", fgColor="F3F4F6")
            lc.border = b
            ws.row_dimensions[row].height = 18

            for col in range(2, total_cols + 1):
                ec = ws.cell(row=row, column=col)
                ec.fill = PatternFill("solid", fgColor="FFFFFF")
                ec.border = b

        for br in day_breaks:
            col = hall_col.get(br.hall_id)
            if col is None:
                continue
            s = max(0, _time_to_slot(br.start_time))
            e = min(TOTAL_SLOTS, _time_to_slot(br.end_time))
            if s >= e:
                continue
            r1 = GRID_FIRST_ROW + s
            r2 = GRID_FIRST_ROW + e - 1
            if r2 > r1:
                ws.merge_cells(start_row=r1, start_column=col, end_row=r2, end_column=col)
            cell = ws.cell(row=r1, column=col)
            cell.value = f"Перерыв\n{br.start_time.strftime('%H:%M')}–{br.end_time.strftime('%H:%M')}"
            cell.fill = PatternFill("solid", fgColor="FEF3C7")
            cell.font = Font(bold=True, color="92400E", size=9)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            amber = Side(style="medium", color="D97706")
            cell.border = Border(left=amber, right=amber, top=amber, bottom=amber)

        KEYNOTE_BG, KEYNOTE_FG = "1E1B4B", "FFFFFF"
        BROADCAST_BG, BROADCAST_FG = "312E81", "C7D2FE"

        for talk in sorted(day_talks, key=lambda t: (0 if t.get("is_keynote") else 1)):
            col = hall_col.get(talk["hall_id"])
            if col is None:
                continue
            s = max(0, _time_to_slot(talk["start"]))
            e = min(TOTAL_SLOTS, _time_to_slot(talk["end"]))
            if s >= e:
                continue
            r1 = GRID_FIRST_ROW + s
            r2 = GRID_FIRST_ROW + e - 1

            ti = talk["track_idx"]
            bg = TRACK_BG[ti % len(TRACK_BG)] if ti >= 0 else "DBEAFE"
            fg = TRACK_FG[ti % len(TRACK_FG)] if ti >= 0 else "1E3A8A"
            time_str = f"{talk['start'].strftime('%H:%M')}–{talk['end'].strftime('%H:%M')}"

            if talk.get("is_keynote"):
                parts = [f"KEYNOTE: {talk['title']}"]
                if talk["speaker"]:
                    parts.append(talk["speaker"])
                parts.append(f"🎤 Выступление  {time_str}")
                if r2 > r1:
                    ws.merge_cells(start_row=r1, start_column=col, end_row=r2, end_column=col)
                cell = ws.cell(row=r1, column=col)
                cell.value = "\n".join(parts)
                cell.fill = PatternFill("solid", fgColor=KEYNOTE_BG)
                cell.font = Font(color=KEYNOTE_FG, size=9, bold=True)
                cell.alignment = Alignment(vertical="center", horizontal="center", wrap_text=True)
                gold = Side(style="medium", color="F59E0B")
                cell.border = Border(left=gold, right=gold, top=gold, bottom=gold)

                broadcast_text = f"📺 Трансляция\n{talk['title']}\n{time_str}"
                for h in halls:
                    bc = hall_col[h.id]
                    if bc == col:
                        continue
                    if r2 > r1:
                        ws.merge_cells(start_row=r1, start_column=bc, end_row=r2, end_column=bc)
                    bcell = ws.cell(row=r1, column=bc)
                    bcell.value = broadcast_text
                    bcell.fill = PatternFill("solid", fgColor=BROADCAST_BG)
                    bcell.font = Font(color=BROADCAST_FG, size=9, italic=True)
                    bcell.alignment = Alignment(vertical="center", horizontal="center", wrap_text=True)
                    dashed = Side(style="dashed", color="6366F1")
                    bcell.border = Border(left=dashed, right=dashed, top=dashed, bottom=dashed)
            else:
                parts = [talk["title"]]
                if talk["speaker"]:
                    parts.append(talk["speaker"])
                if talk["track"]:
                    parts.append(f"[{talk['track']}]")
                parts.append(time_str)
                if r2 > r1:
                    ws.merge_cells(start_row=r1, start_column=col, end_row=r2, end_column=col)
                cell = ws.cell(row=r1, column=col)
                cell.value = "\n".join(parts)
                cell.fill = PatternFill("solid", fgColor=bg)
                cell.font = Font(color=fg, size=9)
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                accent = Side(style="medium", color=fg)
                cell.border = Border(left=accent, right=thin_side, top=thin_side, bottom=thin_side)

        current_row = GRID_FIRST_ROW + TOTAL_SLOTS
        if day_idx < len(sorted_days) - 1:
            ws.row_dimensions[current_row].height = 12
            current_row += 1

    if current_row == 1:
        ws.cell(row=1, column=1, value="Нет распределённых докладов")

    ws.freeze_panes = "B3"

    if tracks:
        ws_legend = wb.create_sheet("Треки")
        ws_legend.column_dimensions["A"].width = 4
        ws_legend.column_dimensions["B"].width = 32
        ws_legend.column_dimensions["C"].width = 18

        for col_idx, header in enumerate(["", "Трек", "Цвет"], start=1):
            hc = ws_legend.cell(row=1, column=col_idx, value=header)
            hc.font = Font(bold=True, color="FFFFFF", size=10)
            hc.fill = PatternFill("solid", fgColor="4F46E5")
            hc.alignment = Alignment(horizontal="center", vertical="center")
            hc.border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
        ws_legend.row_dimensions[1].height = 24

        for i, track in enumerate(tracks):
            row = i + 2
            bg = TRACK_BG[i % len(TRACK_BG)]
            fg = TRACK_FG[i % len(TRACK_FG)]

            dot_cell = ws_legend.cell(row=row, column=1)
            dot_cell.fill = PatternFill("solid", fgColor=bg)
            dot_cell.border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

            name_cell = ws_legend.cell(row=row, column=2, value=track.name)
            name_cell.font = Font(color=fg, bold=True, size=10)
            name_cell.fill = PatternFill("solid", fgColor=bg)
            name_cell.alignment = Alignment(vertical="center", indent=1)
            name_cell.border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

            sample_cell = ws_legend.cell(row=row, column=3, value="Пример доклада")
            sample_cell.font = Font(color=fg, size=9)
            sample_cell.fill = PatternFill("solid", fgColor=bg)
            sample_cell.alignment = Alignment(horizontal="center", vertical="center")
            accent = Side(style="medium", color=fg)
            sample_cell.border = Border(left=accent, right=thin_side, top=thin_side, bottom=thin_side)
            ws_legend.row_dimensions[row].height = 22

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    conf_slug = conference.name.replace(" ", "_")[:30]
    ver_slug = f"_v{version_id}" if version_id else "_draft"
    filename = f"schedule_{conf_slug}{ver_slug}.xlsx"

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
